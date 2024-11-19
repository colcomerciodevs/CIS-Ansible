#!/usr/bin/env python3

import sys
import json
from openpyxl import load_workbook

# Ruta del archivo Excel
excel_file = 'inventory_data.xlsx'

def parse_excel(file):
    try:
        # Cargar el archivo Excel
        wb = load_workbook(file)
        sheet = wb.active  # Seleccionar la hoja activa
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo '{file}'")
        sys.exit(1)
    except Exception as e:
        print(f"Error al leer el archivo Excel: {e}")
        sys.exit(1)

    # Estructura del inventario
    inventory = {"_meta": {"hostvars": {}}}

    try:
        # Iterar por las filas del archivo Excel (omitiendo la cabecera)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            host, group, ip, user = row  # Extraer columnas según el orden
            if not host or not group or not ip or not user:
                print("Error: Una de las filas tiene valores vacíos. Verifica el archivo Excel.")
                sys.exit(1)

            # Agregar host al grupo correspondiente
            if group not in inventory:
                inventory[group] = {"hosts": [], "vars": {}}
            inventory[group]["hosts"].append(host)

            # Agregar variables del host
            inventory["_meta"]["hostvars"][host] = {
                "ansible_host": ip,
                "ansible_user": user
            }

    except Exception as e:
        print(f"Error al procesar el archivo Excel: {e}")
        sys.exit(1)

    return inventory

def main():
    if len(sys.argv) == 2 and sys.argv[1] == '--list':
        # Generar inventario dinámico
        inventory = parse_excel(excel_file)
        print(json.dumps(inventory, indent=2))
    elif len(sys.argv) == 3 and sys.argv[1] == '--host':
        # Devuelve las variables específicas de un host (opcional)
        host = sys.argv[2]
        print(json.dumps({}))
    else:
        print("Uso: --list | --host <nombre_host>")
        sys.exit(1)

if __name__ == '__main__':
    main()
